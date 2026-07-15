"""XLSX 解析: openpyxl 读全部 sheet → markdown 文本。

每个 sheet 渲染为一个 markdown 表。空 sheet 跳过。

注意: openpyxl 3.1.5 解析某些 xlsx 的 styles.xml 时报
"Fill() takes no arguments" 错误 (空 <fill/> 节点), 先用 _sanitize_xlsx 修复。
"""
from __future__ import annotations

import os
import re
import tempfile
import zipfile
from pathlib import Path
from typing import Union

from openpyxl import load_workbook

PathLike = Union[str, Path]


def _sanitize_xlsx(src: Path) -> Path:
    """修复 <fill/> 空 fill 节点, 让 openpyxl 能解析。返回新文件路径 (tmp)。

    openpyxl 在解析 styles.xml 时遇到 `<fill/>` 会尝试构造 PatternFill() 但
    Fill 抽象基类不接受 None 参数。补丁: 把 `<fill/>` 替换成
    `<fill><patternFill patternType="none"/></fill>`。
    """
    src = Path(src)
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", prefix="aog_xlsx_")
    os.close(fd)
    try:
        with zipfile.ZipFile(src) as zin:
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == "xl/styles.xml":
                        try:
                            txt = data.decode("utf-8")
                        except UnicodeDecodeError:
                            txt = data.decode("utf-8", errors="replace")
                        txt = re.sub(
                            r"<fill\s*/>",
                            '<fill><patternFill patternType="none"/></fill>',
                            txt,
                        )
                        txt = re.sub(
                            r"<fill>\s*</fill>",
                            '<fill><patternFill patternType="none"/></fill>',
                            txt,
                        )
                        data = txt.encode("utf-8")
                    zout.writestr(item, data)
    except Exception:
        # sanitize 失败时返回原文件, 让 openpyxl 自己处理 (可能仍报错)
        if os.path.exists(tmp):
            os.unlink(tmp)
        return src
    return Path(tmp)


def _open_xlsx_safe(path: Path):
    """打开 xlsx, 先 sanitize 一次; sanitize 失败或仍报错时回退到原文件。"""
    try:
        sanitized = _sanitize_xlsx(path)
        wb = load_workbook(str(sanitized), data_only=True, read_only=True)
        return wb, sanitized
    except Exception:
        # 仍失败, 直接试原文件
        return load_workbook(str(path), data_only=True, read_only=True), path


def parse_xlsx(path: PathLike) -> str:
    """读 .xlsx → markdown 字符串。

    多 sheet 用 ## sheet_name 切分, 每个 sheet 一个表格。
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"xlsx 文件不存在: {p}")
    if p.suffix.lower() != ".xlsx":
        raise ValueError(f"不是 .xlsx 文件: {p}")

    wb, _tmp = _open_xlsx_safe(p)
    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                row_vals = ["" if c is None else str(c).strip() for c in row]
                if any(v for v in row_vals):
                    rows.append(row_vals)

            if not rows:
                continue

            parts.append(f"## {sheet_name}")
            parts.append("")
            header = rows[0]
            parts.append("| " + " | ".join(header) + " |")
            parts.append("| " + " | ".join(["---"] * len(header)) + " |")
            for row in rows[1:]:
                while len(row) < len(header):
                    row.append("")
                parts.append("| " + " | ".join(v.replace("|", "\\|").replace("\n", " ") for v in row) + " |")
            parts.append("")
    finally:
        wb.close()
        # 清理 tmp
        if str(_tmp) != str(p) and _tmp.exists():
            try:
                _tmp.unlink()
            except OSError:
                pass

    return "\n".join(parts).strip()


def parse_xlsx_first_sheet_rows(path: PathLike, max_rows: int = 50) -> list[list[str]]:
    """读 xlsx 第一个 sheet 前 N 行, 用于快速抽取 (e.g. parts 表格)。"""
    p = Path(path)
    wb, _tmp = _open_xlsx_safe(p)
    try:
        ws = wb[wb.sheetnames[0]]
        rows: list[list[str]] = []
        for row in ws.iter_rows(max_row=max_rows, values_only=True):
            row_vals = ["" if c is None else str(c).strip() for c in row]
            if any(v for v in row_vals):
                rows.append(row_vals)
    finally:
        wb.close()
        if str(_tmp) != str(p) and _tmp.exists():
            try:
                _tmp.unlink()
            except OSError:
                pass
    return rows
